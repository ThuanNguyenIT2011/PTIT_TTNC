from __future__ import annotations

from pathlib import Path
from typing import Union

import csv
from io import StringIO

from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook


class DocumentReader:
    TEXT_EXTENSIONS = {
        ".txt", ".md", ".log", ".ini", ".cfg", ".conf",
        ".json", ".xml", ".yaml", ".yml", ".csv"
    }

    CODE_EXTENSIONS = {
        ".py", ".js", ".ts", ".tsx", ".jsx",
        ".java", ".c", ".cpp", ".h", ".hpp",
        ".cs", ".php", ".rb", ".go", ".rs",
        ".swift", ".kt", ".scala", ".sh", ".bat",
        ".ps1", ".sql", ".html", ".htm", ".css"
    }

    OFFICE_EXTENSIONS = {
        ".docx", ".pdf", ".xlsx"
    }

    SUPPORTED_EXTENSIONS = TEXT_EXTENSIONS | CODE_EXTENSIONS | OFFICE_EXTENSIONS

    def read(self, file_path: Union[str, Path], encoding: str = "utf-8") -> str:
        path = Path(file_path)

        if not path.exists():
            raise FileNotFoundError(f"Không tìm thấy file: {path}")

        if not path.is_file():
            raise ValueError(f"Đây không phải file hợp lệ: {path}")

        ext = path.suffix.lower()

        if ext not in self.SUPPORTED_EXTENSIONS:
            raise ValueError(
                f"Định dạng không hỗ trợ: {ext}. "
                f"Chỉ hỗ trợ: {', '.join(sorted(self.SUPPORTED_EXTENSIONS))}"
            )

        if ext == ".pdf":
            return self._read_pdf(path)

        if ext == ".docx":
            return self._read_docx(path)

        if ext == ".xlsx":
            return self._read_xlsx(path)

        if ext == ".csv":
            return self._read_csv(path, encoding)

        if ext in self.TEXT_EXTENSIONS or ext in self.CODE_EXTENSIONS:
            return self._read_text_file(path, encoding)

        raise ValueError(f"Không thể xử lý file: {path}")

    def _read_text_file(self, path: Path, encoding: str = "utf-8") -> str:
        encodings_to_try = [encoding, "utf-8", "utf-8-sig", "cp1252", "latin-1"]

        for enc in encodings_to_try:
            try:
                return path.read_text(encoding=enc)
            except UnicodeDecodeError:
                continue

        return path.read_text(encoding="utf-8", errors="ignore")

    def _read_csv(self, path: Path, encoding: str = "utf-8") -> str:
        content = self._read_text_file(path, encoding)

        try:
            sample = content[:2048]
            dialect = csv.Sniffer().sniff(sample)
            reader = csv.reader(StringIO(content), dialect)
        except Exception:
            reader = csv.reader(StringIO(content))

        rows = []
        for row in reader:
            rows.append(" | ".join(str(cell) for cell in row))

        return "\n".join(rows).strip()

    def _read_pdf(self, path: Path) -> str:
        reader = PdfReader(str(path))
        pages_text = []

        for page_index, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            if text.strip():
                pages_text.append(f"--- Page {page_index} ---\n{text}")

        return "\n\n".join(pages_text).strip()

    def _read_docx(self, path: Path) -> str:
        doc = Document(str(path))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n".join(paragraphs).strip()

    def _read_xlsx(self, path: Path) -> str:
        workbook = load_workbook(filename=str(path), data_only=True)
        parts = []

        for sheet in workbook.worksheets:
            parts.append(f"=== Sheet: {sheet.title} ===")

            for row in sheet.iter_rows(values_only=True):
                row_values = ["" if cell is None else str(cell) for cell in row]
                if any(value.strip() for value in row_values):
                    parts.append(" | ".join(row_values))

            parts.append("")

        return "\n".join(parts).strip()


if __name__ == "__main__":
    reader = DocumentReader()

    file_path = "./Main.py"  # đổi thành file của bạn
    try:
        content = reader.read(file_path)
        print(content)
    except Exception as e:
        print(f"Lỗi: {e}")